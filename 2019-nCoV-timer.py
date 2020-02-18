import os
import json
import requests
import datetime
import warnings
import schedule
from time import sleep
from openpyxl import load_workbook, Workbook
warnings.filterwarnings('ignore')

                     #  省份           城市,  确诊, 死亡,  治愈, 疑似
# China = [{"province":"河南", cities:[[信阳", 10,   0,    10,   10], [], []]}, {}, {}, {}]
China = []

def getData():
    url = "https://lab.isaaclin.cn/nCoV/api/area"
    headers = {}
    headers['user-agent'] = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.77 Safari/537.36' #http头大小写不敏感
    headers['accept'] = 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8'
    headers['Connection'] = 'keep-alive'
    headers['Upgrade-Insecure-Requests'] = '1'
    
    while True:
        r = requests.get(url, headers=headers)
        r.encoding = r.apparent_encoding
        json_str = json.loads(r.text)
        if json_str['success']:
            break
    
    for data in json_str['results']:
        if data['countryName'] == '中国': 
            province_dict = {}
            province_dict["province"] = data['provinceShortName']
            province_dict["cities"] = []
            if len(data['cities']) == 0:
                province_dict['cities'].append(get_city_data('', data["confirmedCount"], data['deadCount'], data['curedCount'], data['suspectedCount']))            
            else:
                for city in data['cities']:        
                    if city['cityName'] != '待明确地区':
                        province_dict['cities'].append(get_city_data(city["cityName"], city["confirmedCount"], city['deadCount'], city['curedCount'], city['suspectedCount']))
            
            China.append(province_dict)

def get_city_data(name, confirmedCount, deadCount, curedCount, suspectedCount):
    tmp = []
    tmp.append(name)
    tmp.append(confirmedCount)
    tmp.append(deadCount)
    tmp.append(curedCount)
    tmp.append(suspectedCount)
    return tmp


def data_to_excel(my_sheet, cur_day):
    row = ['省份', '城市', '日期', '确诊', '死亡', '治愈', '疑似']
    my_sheet.append(row)
    for i in range(len(China)):
        province = China[i]['province']
        for city in China[i]['cities']:
            tmp = []
            tmp = [province, city[0], cur_day, city[1], city[2], city[3], city[4]]
            my_sheet.append(tmp)

def main():

    cur_day = datetime.date.today().strftime('%m-%d')
    try:
        getData()
    except:
        print("接口出错")

    fname = "2019-nCoV.xlsx"

    try:
        if os.path.exists(fname):
            wb = load_workbook(fname)
        else:
            wb = Workbook()
        my_sheet = wb.create_sheet(cur_day)
    except:
        print("文件路径有问题，请检查是否打开了文件，如果打开了，请关闭")
    
    try:
        data_to_excel(my_sheet, cur_day)
        cur_day = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(cur_day + " 文件写入成功")
    except:
        print("文件写入过程中出现问题")
    finally:
        wb.save(fname)


if __name__ == '__main__':
    schedule.every().day.at("00:00").do(main)
    while True:
        schedule.run_pending()
        sleep(10)